"""
Generate a professional single-elimination 16-player tournament bracket
template in Excel (.xlsx).

Layout (landscape, A3-like proportions):
  Col A      : narrow spacer / seed column (dark navy)
  Cols B-F   : Round of 16  player names  (2 per match × 8 matches)
  Col G      : bracket connector / junction column
  Cols H-L   : Quarterfinals             (2 per match × 4 matches)
  Col M      : bracket connector
  Cols N-R   : Semifinals                (2 per match × 2 matches)
  Col S      : bracket connector
  Cols T-X   : Final                     (2 slots × 1 match)
  Col Y      : champion column

Row structure per match (4 rows):
  row+0  player 1  (seed | name | arm right+bottom)
  row+1  spacer    (seed | blank | arm top+bottom)
  row+2  player 2  (seed | name | arm top+right)
  row+3  junction  (navy | blank | junction box)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup

# ─── Palette ────────────────────────────────────────────────────────────────
NAVY      = "17305E"
WHITE     = "FFFFFF"
LIGHT_BG  = "F4F6FA"   # very subtle tint for player cells
HDR_BG    = "17305E"
JUNC_BG   = "17305E"
CHAMP_BG  = "155724"   # dark green champion
ACC_GOLD  = "C9A84C"   # accent gold for title underline
GRAY_LINE = "1A3A6B"   # bracket line colour (dark navy for crisp lines)
SUBTITLE_FG = "445577"

# ─── Fill helpers ────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

navy_fill  = fill(NAVY)
white_fill = fill(WHITE)
light_fill = fill(LIGHT_BG)
junc_fill  = fill(JUNC_BG)
champ_fill = fill(CHAMP_BG)
empty_fill = fill(WHITE)

# ─── Side / Border helpers ───────────────────────────────────────────────────
def side(style="thin", color=GRAY_LINE):
    return Side(style=style, color=color)

NONE = Side(style=None)
TH   = side("thin")
MD   = side("medium")
TK   = side("thick")

def border(top=NONE, right=NONE, bottom=NONE, left=NONE):
    return Border(top=top, right=right, bottom=bottom, left=left)

ALL_MD   = border(MD, MD, MD, MD)
ALL_TH   = border(TH, TH, TH, TH)

# arm borders
ARM_P1   = border(top=NONE,  right=MD, bottom=MD, left=NONE)   # right + bottom
ARM_SP   = border(top=MD,    right=NONE, bottom=MD, left=NONE)  # top + bottom (spine)
ARM_P2   = border(top=MD,    right=MD, bottom=NONE, left=NONE)  # top + right

# player name cell: top + bottom + left (open right toward arm)
P_BORDER = border(top=MD, right=NONE, bottom=MD, left=MD)
P_BORDER_TH = border(top=TH, right=NONE, bottom=TH, left=TH)

# ─── Font helpers ────────────────────────────────────────────────────────────
def font(bold=False, sz=10, color=None, italic=False):
    kw = dict(bold=bold, size=sz, italic=italic)
    if color:
        kw["color"] = color
    return Font(**kw)

# ─── Alignment helpers ───────────────────────────────────────────────────────
CTR = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_CTR = Alignment(horizontal="left", vertical="center", wrap_text=False)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ─── Workbook setup ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Bracket Template"

# Landscape A3-ish
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize   = ws.PAPERSIZE_A3
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 1
ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                               header=0.3, footer=0.3)
ws.print_options.horizontalCentered = True
ws.sheet_view.showGridLines = False

# ─── Column layout ────────────────────────────────────────────────────────────
# Column index (1-based):
#  1(A)   : seed / left margin  (narrow)
#  2-5    : R16 player name (merged 4 wide)
#  6      : R16 arm/junction
#  7-10   : QF player name
#  11     : QF arm
#  12-15  : SF player name
#  16     : SF arm
#  17-20  : Final player name
#  21     : Final arm / champion arrow
#  22-25  : CHAMPION cell
#  26(Z)  : right margin

SEED_COL  = 1    # A
R16_NC    = 2    # B (name start)
R16_NW    = 4    # columns wide for name
R16_ARM   = 6    # F
QF_NC     = 7    # G
QF_NW     = 4
QF_ARM    = 11   # K
SF_NC     = 12   # L
SF_NW     = 4
SF_ARM    = 16   # P
FIN_NC    = 17   # Q
FIN_NW    = 4
FIN_ARM   = 21   # U
CHAMP_NC  = 22   # V
CHAMP_NW  = 4
TOTAL_COLS = 26

# Column widths (units)
col_widths = {
    1:  2.5,   # seed
    2:  3.0,   3:  3.0,   4:  3.0,   5:  8.5,  # R16 name (B-E wide, last carries text)
    6:  2.0,   # R16 arm
    7:  3.0,   8:  3.0,   9:  3.0,   10: 8.5,  # QF name
    11: 2.0,   # QF arm
    12: 3.0,   13: 3.0,   14: 3.0,   15: 8.5,  # SF name
    16: 2.0,   # SF arm
    17: 3.0,   18: 3.0,   19: 3.0,   20: 8.5,  # Final name
    21: 2.0,   # Final arm
    22: 3.0,   23: 3.0,   24: 3.0,   25: 10.0, # Champion
    26: 1.5,   # right margin
}
for col_idx, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# ─── Helper: apply style to a cell ────────────────────────────────────────────
def style_cell(ws, row, col, value="", fnt=None, fll=None, bdr=None, aln=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fnt: cell.font      = fnt
    if fll: cell.fill      = fll
    if bdr: cell.border    = bdr
    if aln: cell.alignment = aln
    return cell

def merge_style(ws, r1, c1, r2, c2, value="", fnt=None, fll=None, bdr=None, aln=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return style_cell(ws, r1, c1, value, fnt, fll, bdr, aln)

# ─── Row layout constants ─────────────────────────────────────────────────────
HEADER_ROWS = 5   # rows 1-5: title block
ROWS_PER_MATCH = 4
NUM_R16_MATCHES = 8
DATA_ROWS = NUM_R16_MATCHES * ROWS_PER_MATCH   # 32 rows
FOOTER_ROW_START = HEADER_ROWS + DATA_ROWS + 2

TOTAL_ROWS = FOOTER_ROW_START + 5

# Set row heights
ws.row_dimensions[1].height = 36   # main title
ws.row_dimensions[2].height = 20   # subtitle
ws.row_dimensions[3].height = 14   # gold accent line
ws.row_dimensions[4].height = 18   # column headers
ws.row_dimensions[5].height = 6    # gap below header

for mi in range(NUM_R16_MATCHES):
    base = HEADER_ROWS + mi * ROWS_PER_MATCH
    ws.row_dimensions[base + 1].height = 18   # P1
    ws.row_dimensions[base + 2].height = 7    # spacer
    ws.row_dimensions[base + 3].height = 18   # P2
    ws.row_dimensions[base + 4].height = 8    # junction gap

ws.row_dimensions[FOOTER_ROW_START    ].height = 8
ws.row_dimensions[FOOTER_ROW_START + 1].height = 14
ws.row_dimensions[FOOTER_ROW_START + 2].height = 12
ws.row_dimensions[FOOTER_ROW_START + 3].height = 12
ws.row_dimensions[FOOTER_ROW_START + 4].height = 12

# ─── TITLE BLOCK (rows 1-5) ───────────────────────────────────────────────────
# Row 1: dark navy header bar spanning full width
merge_style(ws, 1, 1, 1, TOTAL_COLS,
    "TAEKWONDO CHAMPIONSHIP 2025",
    fnt=Font(bold=True, size=20, color=WHITE, name="Calibri"),
    fll=fill(NAVY),
    aln=Alignment(horizontal="center", vertical="center"))

# Row 2: subtitle
merge_style(ws, 2, 1, 2, TOTAL_COLS,
    "SENIOR MEN  ·  -68KG",
    fnt=Font(bold=False, size=11, color="AABBCC", name="Calibri"),
    fll=fill(NAVY),
    aln=Alignment(horizontal="center", vertical="center"))

# Row 3: thin gold accent separator
merge_style(ws, 3, 1, 3, TOTAL_COLS,
    "",
    fll=fill(ACC_GOLD),
    bdr=border())

# Row 4: column round headers
ROUND_HEADERS = [
    (R16_NC,  R16_NC + R16_NW - 1,  "ROUND OF 16"),
    (QF_NC,   QF_NC  + QF_NW  - 1,  "QUARTERFINALS"),
    (SF_NC,   SF_NC  + SF_NW  - 1,  "SEMIFINALS"),
    (FIN_NC,  FIN_NC + FIN_NW - 1,  "FINAL"),
    (CHAMP_NC, CHAMP_NC + CHAMP_NW - 1, "CHAMPION"),
]
# fill full header row navy
for c in range(1, TOTAL_COLS + 1):
    style_cell(ws, 4, c, fll=fill(NAVY))
for (c1, c2, lbl) in ROUND_HEADERS:
    merge_style(ws, 4, c1, 4, c2,
        lbl,
        fnt=Font(bold=True, size=9, color=WHITE, name="Calibri"),
        fll=fill(NAVY),
        bdr=border(left=side("medium", WHITE), right=side("medium", WHITE)),
        aln=CTR)
# seed col header
style_cell(ws, 4, SEED_COL, "#",
    fnt=Font(bold=True, size=8, color=WHITE),
    fll=fill(NAVY), aln=CTR)

# Row 5: tiny gap — fill navy to match header bottom
for c in range(1, TOTAL_COLS + 1):
    style_cell(ws, 5, c, fll=fill(WHITE))

# ─── BRACKET DATA ROWS ────────────────────────────────────────────────────────
# Seed numbers follow standard seeding order for 16-player bracket:
# Match 1: 1 vs 16,  Match 2: 9 vs 8,  Match 3: 5 vs 12,  Match 4: 13 vs 4
# Match 5: 3 vs 14,  Match 6: 11 vs 6, Match 7: 7 vs 10,  Match 8: 15 vs 2
SEEDS = [
    (1,  16), (9,  8),
    (5,  12), (13, 4),
    (3,  14), (11, 6),
    (7,  10), (15, 2),
]

def player_label(seed, round_num=1):
    if round_num == 1:
        return f"Player {seed}"
    return ""

# ─── Round of 16 ─────────────────────────────────────────────────────────────
for mi in range(NUM_R16_MATCHES):
    base  = HEADER_ROWS + mi * ROWS_PER_MATCH
    rowP1 = base + 1
    rowSp = base + 2
    rowP2 = base + 3
    rowGp = base + 4
    s1, s2 = SEEDS[mi]

    # -- Seed column --
    style_cell(ws, rowP1, SEED_COL, s1,
        fnt=Font(bold=True, size=8, color=WHITE, name="Calibri"),
        fll=navy_fill, aln=CTR,
        bdr=ALL_TH)
    style_cell(ws, rowSp, SEED_COL, "",
        fll=navy_fill, bdr=border(left=TH, right=TH))
    style_cell(ws, rowP2, SEED_COL, s2,
        fnt=Font(bold=True, size=8, color=WHITE, name="Calibri"),
        fll=navy_fill, aln=CTR,
        bdr=ALL_TH)
    style_cell(ws, rowGp, SEED_COL, "",
        fll=navy_fill, bdr=border(left=TH, right=TH, bottom=TH))

    # -- Player name cells (merged B:E) --
    # P1 row
    merge_style(ws, rowP1, R16_NC, rowP1, R16_NC + R16_NW - 1,
        f"Player {s1}",
        fnt=Font(size=9, color="111111", name="Calibri"),
        fll=light_fill,
        bdr=P_BORDER,
        aln=LEFT_CTR)
    # spacer
    merge_style(ws, rowSp, R16_NC, rowSp, R16_NC + R16_NW - 1, "",
        fll=white_fill)
    # P2 row
    merge_style(ws, rowP2, R16_NC, rowP2, R16_NC + R16_NW - 1,
        f"Player {s2}",
        fnt=Font(size=9, color="111111", name="Calibri"),
        fll=light_fill,
        bdr=P_BORDER,
        aln=LEFT_CTR)
    # gap
    merge_style(ws, rowGp, R16_NC, rowGp, R16_NC + R16_NW - 1, "",
        fll=white_fill)

    # -- Arm/junction column (F) --
    style_cell(ws, rowP1, R16_ARM, "", fll=white_fill, bdr=ARM_P1)
    style_cell(ws, rowSp, R16_ARM, "", fll=white_fill, bdr=ARM_SP)
    style_cell(ws, rowP2, R16_ARM, "", fll=white_fill, bdr=ARM_P2)
    # Junction box (small navy square)
    style_cell(ws, rowGp, R16_ARM, "",
        fll=junc_fill, bdr=ALL_MD)

# ─── LATER ROUNDS helper ──────────────────────────────────────────────────────
# Each subsequent round: each match spans 2 feeder matches = 8 rows
# For QF: 4 matches each spanning 8 rows (rows 0-7 from base)
# For SF: 2 matches each spanning 16 rows
# For Final: 1 match spanning 32 rows

def draw_later_round(ws, round_idx, name_col, name_width, arm_col,
                     num_matches, rows_per_match_span, is_final=False):
    """
    round_idx       : 0=R16(skip), 1=QF, 2=SF, 3=Final
    name_col        : starting column for name cells
    name_width      : how many columns merged for name
    arm_col         : arm/junction column
    num_matches     : number of matches in this round
    rows_per_match  : total rows this round's single match spans
    """
    for mi in range(num_matches):
        top = HEADER_ROWS + mi * rows_per_match_span + 1  # 1-based
        bot = top + rows_per_match_span - 1

        mid = (top + bot) // 2  # vertical midpoint for the junction box row

        # Player name cell (merged top to bot)
        merge_style(ws, top, name_col, bot, name_col + name_width - 1,
            "",
            fnt=Font(size=9, bold=True, color="111111", name="Calibri"),
            fll=light_fill if not is_final else fill("EAF4FF"),
            bdr=P_BORDER,
            aln=LEFT_CTR)

        if not is_final:
            # Arm/junction column: spine from top to junction box at bottom
            for r in range(top, bot + 1):
                if r == top:
                    style_cell(ws, r, arm_col, "", fll=white_fill, bdr=ARM_P1)
                elif r == bot:
                    style_cell(ws, r, arm_col, "", fll=junc_fill,  bdr=ALL_MD)
                elif r <= mid:
                    style_cell(ws, r, arm_col, "", fll=white_fill, bdr=ARM_SP)
                elif r == mid + 1:
                    style_cell(ws, r, arm_col, "", fll=white_fill, bdr=ARM_P2)
                else:
                    style_cell(ws, r, arm_col, "", fll=white_fill, bdr=ARM_SP)

# QF: 4 matches, each spans 8 rows (2 R16 matches × 4 rows each)
draw_later_round(ws, 1, QF_NC, QF_NW, QF_ARM,
                 num_matches=4, rows_per_match_span=8)

# SF: 2 matches, each spans 16 rows
draw_later_round(ws, 2, SF_NC, SF_NW, SF_ARM,
                 num_matches=2, rows_per_match_span=16)

# Final: 1 match spans 32 rows (top half for player 1, bottom for player 2)
# We'll draw two halves manually for the final (two separate name cells)
FIN_TOP    = HEADER_ROWS + 1
FIN_TOTAL  = NUM_R16_MATCHES * ROWS_PER_MATCH   # 32
FIN_MID    = FIN_TOP + FIN_TOTAL // 2 - 1       # end of top half

# Final P1 (top half, rows 6-21)
merge_style(ws, FIN_TOP, FIN_NC, FIN_MID, FIN_NC + FIN_NW - 1,
    "",
    fnt=Font(size=10, bold=True, color="111111", name="Calibri"),
    fll=light_fill,
    bdr=P_BORDER,
    aln=LEFT_CTR)

# Final P2 (bottom half, rows 22-37)
merge_style(ws, FIN_MID + 1, FIN_NC, FIN_TOP + FIN_TOTAL - 1,
            FIN_NC + FIN_NW - 1,
    "",
    fnt=Font(size=10, bold=True, color="111111", name="Calibri"),
    fll=light_fill,
    bdr=P_BORDER,
    aln=LEFT_CTR)

# Final arm: junction at absolute middle
for r in range(FIN_TOP, FIN_TOP + FIN_TOTAL):
    mid_final = FIN_TOP + FIN_TOTAL // 2 - 1
    if r == FIN_TOP:
        style_cell(ws, r, FIN_ARM, "", fll=white_fill, bdr=ARM_P1)
    elif r == mid_final:
        style_cell(ws, r, FIN_ARM, "", fll=junc_fill, bdr=ALL_MD)
    elif r < mid_final:
        style_cell(ws, r, FIN_ARM, "", fll=white_fill, bdr=ARM_SP)
    elif r == mid_final + 1:
        style_cell(ws, r, FIN_ARM, "", fll=white_fill, bdr=ARM_P2)
    else:
        style_cell(ws, r, FIN_ARM, "", fll=white_fill, bdr=ARM_SP)

# ─── CHAMPION cell ────────────────────────────────────────────────────────────
CHAMP_TOP = HEADER_ROWS + FIN_TOTAL // 4       # vertically centred
CHAMP_BOT = HEADER_ROWS + FIN_TOTAL * 3 // 4

merge_style(ws, CHAMP_TOP, CHAMP_NC, CHAMP_BOT, CHAMP_NC + CHAMP_NW - 1,
    "🏆  CHAMPION",
    fnt=Font(bold=True, size=12, color=WHITE, name="Calibri"),
    fll=fill(CHAMP_BG),
    bdr=border(TK, TK, TK, TK),
    aln=CTR)

# fill surrounding champion area white
for r in range(HEADER_ROWS + 1, HEADER_ROWS + FIN_TOTAL + 1):
    for c in range(CHAMP_NC, CHAMP_NC + CHAMP_NW + 1):
        cell = ws.cell(row=r, column=c)
        if cell.value is None and not getattr(cell, '_merged', False):
            cell.fill = white_fill

# ─── FOOTER ──────────────────────────────────────────────────────────────────
fr = FOOTER_ROW_START

# Thin separator line
merge_style(ws, fr, 1, fr, TOTAL_COLS, "",
    fll=fill(ACC_GOLD))

# Notes
merge_style(ws, fr + 1, 1, fr + 1, TOTAL_COLS,
    "NOTES:",
    fnt=Font(bold=True, size=8, color=NAVY, name="Calibri"),
    aln=Alignment(horizontal="left", vertical="center"))

notes = [
    "• Matches are best of 3 rounds.",
    "• Tie-breaking according to WT Competition Rules.",
    "• This bracket template is editable — type directly into the player name cells.",
]
for i, note in enumerate(notes):
    merge_style(ws, fr + 2 + i, 1, fr + 2 + i, TOTAL_COLS,
        note,
        fnt=Font(size=8, color="555555", name="Calibri"),
        aln=Alignment(horizontal="left", vertical="center"))

# ─── FILL BACKGROUND white for all unused cells ───────────────────────────────
for r in range(1, TOTAL_ROWS + 1):
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=r, column=c)
        if cell.fill.patternType in (None, "none"):
            cell.fill = white_fill

# ─── NAMED RANGE hint (cosmetic only) ────────────────────────────────────────
ws["A1"].comment = None  # placeholder: no comments needed

# ─── PRINT AREA & FREEZE PANES ────────────────────────────────────────────────
ws.print_area = f"A1:{get_column_letter(TOTAL_COLS)}{TOTAL_ROWS}"
ws.freeze_panes = "B6"  # freeze title + header rows, seed column

# ─── SAVE ────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\chira\Desktop\TKD-Championship-Manager\TKD_Bracket_Template.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
